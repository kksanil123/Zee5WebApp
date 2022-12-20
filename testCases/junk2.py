import time
import re
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
import openpyxl


# serv = Service()
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(15)
workbook = openpyxl.Workbook()

try:

    driver.get(r"")

    driver.find_element(By.NAME, 'username').send_keys('')
    driver.find_element(By.NAME, 'password').send_keys('')
    driver.find_element(By.XPATH, '//button[text()="Continue"]').click()
    driver.find_element(By.LINK_TEXT, 'DTC-SVOD').click()
    time.sleep(2)

    driver.find_element(By.NAME, 'test-cases-icon').click()
    driver.find_element(By.XPATH, '//*[contains(@class, "Arrow")]/parent::div').click()
    time.sleep(1)

    driver.find_element(By.XPATH, '//*[text() ="Suites"]').click()
    driver.find_element(By.XPATH, '//div[@class = "TestCaseTreeContainer__controls"]/button[1]').click()
    driver.find_element(By.XPATH, '//div[@class = "SavedFilters__title"]/button[2]').click()

    driver.find_element(By.XPATH, '//span[text()= "Active"]').click()
    driver.find_element(By.XPATH, '//div[text() = "Type"]/following-sibling::label[2]').click()
    driver.find_element(By.XPATH, '//span[text()= "Manual"]').click()
    time.sleep(2)
    driver.find_element(By.XPATH, '(//div[@class= "css-7krynh"])[9]/descendant::input').send_keys('ROKU')
    time.sleep(2)
    driver.find_element(By.XPATH,
                        '(//div[@class= "Select css-2b097c-container"])[9]/child::div[2]//*[text()="ROKU"]').click()
    time.sleep(2)

    driver.find_element(By.XPATH, '//button[text()="Dismiss"]').click()

    suites = driver.find_elements(By.XPATH, '//div[@class ="TreeNodeName"]')

    for suite in suites:
        name = suite.text
        regex = '(?=.*[/:\[])'
        p = re.compile(regex)

        if re.search(p, name):
            workbook.create_sheet(name[:2])
            sheet = workbook[name[:2]]
        else:
            workbook.create_sheet(name)
            sheet = workbook[name]

        headers = ['Test Case Id', 'Summary', 'Description', 'Precondition', 'Scenario', 'Expected Result', 'Apps']
        for i in range(1, 2):
            for j in range(1, 8):
                sheet.cell(i, j).value = headers[j - 1]

        print(suite.text)
        suite.click()
        n = 2
        tcs = driver.find_elements(By.XPATH, '//ul[@class="LoadableTreeGroupView__children"]/child::li')
        for tc in tcs:
            tc.click()
            time.sleep(1)
            xpath = ['//span[@class="TestCaseLayout__id"]',
                     '//div[@class="TestCaseLayout__name"]',
                     '//div[text()="Description"]/parent::h3/following-sibling::div/descendant::*',
                     '//div[@class ="CustomFieldValueGroupedList__field-name" and contains(text(),"Preconditions")]/following-sibling::ul/descendant::div[@class="CustomFieldValueGroupedList__value-name"]',
                     '//pre[@class="Multiline"]',
                     '//div[text() ="Expected result"]/parent::h3/following-sibling::div[@class="PaneSection__content"]/child::div[@class="MarkdownArticle"]',
                     '//div[@class ="CustomFieldValueGroupedList__field-name" and contains(text(),"Brand")]/following-sibling::ul/descendant::div[@class="CustomFieldValueGroupedList__value-name"]']

            data = []
            t_id = driver.find_element(By.XPATH, xpath[0]).text
            data.append(t_id)
            t_summary = driver.find_element(By.XPATH, xpath[1]).text
            t_ind = t_summary.find(" ")
            t_summary = t_summary[t_ind+1:]
            data.append(t_summary)
            t_desc = driver.find_element(By.XPATH, xpath[2]).text
            data.append(t_desc)
            t_precons = []
            precons = driver.find_elements(By.XPATH, xpath[3])
            for precon in precons:
                t_precons.append(precon.text)
            data.append(t_precons)
            t_scenarios = []
            scenarios = driver.find_elements(By.XPATH, xpath[4])
            for sceanrio in scenarios:
                t_scenarios.append(sceanrio.text)
            data.append(t_scenarios)
            t_exp_ress = []
            exp_ress = driver.find_elements(By.XPATH, xpath[5])
            for exp_res in exp_ress:
                t_exp_ress.append(exp_res.text)
            data.append(t_exp_ress)
            t_apps = []
            apps = driver.find_elements(By.XPATH, xpath[6])
            for app in apps:
                t_apps.append(app.text)
            data.append(t_apps)

            for i in range(n, n+1):
                for j in range(1, 8):
                    if type(data[j-1]) is list:
                        a = ''
                        k = 1
                        for st in data[j-1]:
                            a = a + f'{k}. {st} \n'
                            k = k + 1
                        sheet.cell(i, j).value = a
                    else:
                        sheet.cell(i, j).value = data[j - 1]

            n = n+1
            # break
        time.sleep(1)
        suite.click()
        print("***********************************")
        # break

except Exception as e:
    print(e.args)
    print("Driver closed...")


finally:
    workbook.save(r'C:\Users\karla\OneDrive\Documents\RokuTestCases.xlsx')
    driver.close()
    pass




